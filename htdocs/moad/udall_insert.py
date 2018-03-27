#!/usr/bin/env python

import sys
import openpyxl as op
import subprocess
#import mysql.connector as mc

def worksheet(wb,strSheet):
	return wb.get_sheet_by_name(strSheet)

def sqlup(CID,TABLE,FIELD,VAL):
	cmd = "update {} set {} = '{}' where CommentID = '{}'".format(TABLE,FIELD,VAL,CID)


FILE=sys.argv[1];
STUDYVISITNUM=sys.argv[2];

# Open the xlsx file
wb = op.load_workbook(FILE);


# Demographics
ws = worksheet(wb,'002 Demographics');

for row in ws.iter_rows():
	SUBJ=row[1].value

	# demographics.gender
	GENDER=row[2].value

	# demographics.ethnicity
	ETH = row[3].value

	if ETH == 'Not Hispanic or Latino':
		ETH = 'nonhisp'
	elif ETH == 'Hispanic or Latino':
		ETH = 'hisp'
	elif ETH == "Not reported":
		ETH = 'nonrep'
	else:
		print "Error: unexpected ethnicity value?"
		#sys.exit()

	# demographics.race
	AI = row[4].value
	ASIAN = row[5].value
	HAW = row[6].value
	BLK = row[7].value
	WHT = row[8].value
	MRACE = row[9].value
	RACEUNK = row[10].value

	RACE = ''
	if AI == 't':
		RACE = 'american_indian/alaska_native'
	elif ASIAN == 't':
	 	RACE = 'asian'
	elif HAW == 't':
		RACE = 'native_hawaiian/pac_islander'
	elif BLK == 't':
		RACE = 'black'
	elif WHT == 't':
		RACE = 'white'
	elif MRACE == 't':
		RACE = 'multirace'
	elif RACEUNK == 't':
		RACE = 'unknown'


	# demographics.educ_level
	EDUC_LEVEL = row[11].value

	# demographics.employ_stat
	EMPLOY_STAT = row[12].value

	# demographics.employ_stat_other
	EMPLOY_STAT_OTHER = row[13].value

	p = subprocess.Popen(['getpscidfromsubjid.sh', SUBJ], stdout=subprocess.PIPE, stderr=subprocess.PIPE)

	out, err = p.communicate()

	## TO DO:  check if PSCID not returned and throw flag
	PSCID = out.strip()
	#print SUBJ + ' ' + PSCID

# Vitals
ws = worksheet(wb,'006 Vitals');
FIELDS = ['clinical.pls_sit', 'clinical.pls_sta', 'clinical.weight','clinical.height','clinical.bmi']
iRow = 0;
for row in ws.iter_rows():
	if iRow != 0:
		SUBJ=row[1].value

		colnum = 2
		for iCol in range(0,5):
			FIELD = FIELDS[iCol]
			VAL = row[colnum].value
			colnum += 1
			cmd = "update clinical set {} = '{}' where CommentID = '{}'".format(FIELD,VAL,'commentid')
			#print cmd

	iRow += 1



# -- TO DO Behavioral History --

# BP Vitals
ws = worksheet(wb,'006BP Vitals');
for row in ws.iter_rows():

	if iRow != 0:
		SUBJ=row[1].value

		POS = row[2].value
		SYS_BP = row[3].value
		DIA_BP = row[4].value

		if POS == 'Sitting':
			F_END = '_sit'
		else:
			F_END = '_sta'

		cmd = "update clinical set bp_s{} = '{}', bp_d{} = '{}' where CommentID = '{}'".format(F_END,SYS_BP,F_END,DIA_BP,'commentid')
		#print cmd


# HY
ws = worksheet(wb,'010 HandY');
for row in ws.iter_rows():

	if iRow != 0:
		SUBJ=row[1].value
		HY = row[2].value
		cmd = "update motor set hy = '{}' where CommentID = '{}'".format(HY,'commentid')
		#print cmd

# PDQ39
ws = worksheet(wb,'026 PDQ39')
for row in ws.iter_rows():

	if iRow != 0:
		SUBJ=row[1].value

		# Edit for AE?
		for colNum in range(1,41):
			if colNum == 29:
				continue
			elif colNum > 29:
				FIELD = 'i{}'.format(colNum-1)
			else:
				FIELD = 'i{}'.format(colNum)
			VAL = row[colNum+1].value.split(' ')[0]

			cmd = "update pdq set {} = '{}' where CommentID = '{}'".format(FIELD,VAL,'commentid')

		# TO DO :  SCORING

# Epworth
ws = worksheet(wb,'027 Epworth Sleep');
for row in ws.iter_rows():
	if iRow != 0:
		SUBJ=row[1].value

		for colNum in range(1,10):
			FIELD = 'i{}'.format(colNum-1)
			VAL = row[colNum+1].value
			cmd = "update epworth set {} = '{}' where CommentID = '{}'".format(FIELD,VAL,'commentid')
			print cmd

		total = row[10].value
		cmd = "update epworth set total = '{}' where CommentID = '{}'".format(total,'commentid')


# MoCA CHECK THIS
ws = worksheet(wb,'038 MoCA');
for row in ws.iter_rows():
	if iRow != 0:
		SUBJ=row[1].value

		visual = row[2].value
		naming = row[3].value
		attn = row[4].value + row[5].value + row[6].value # ???
		language = row[7].value + row[8].value # ???
		abstraction = row[8].value
		delrecall = row[9].value
		orientation = row[10].value # ???
		total = row[12].value

		cmd = "update cognition set moca_visual = '{}', moca_naming = '{}',moca_attention = '{}', \
		moca_language = '{}',moca_abstraction = '{}', moca_delrecall = '{}', moca_orientation = '{}', \
		moca = '{}' where CommentID='{}'".format(visual,naming,attn,language,abstraction,delrecall,orientation,total,'commentid')

# MDS-UPDRS
ws = worksheet(wb,'011 MDS-UPDRS');
for row in ws.iter_rows():
	if iRow != 0:
		SUBJ=row[1].value
		CID='commentid'
		sqlup(CID,'mdsupdrs',cog,row[3].value)
		sqlup(CID,'mdsupdrs',hal_psy,row[4].value)
		sqlup(CID,'mdsupdrs',dep,row[5].value)
		sqlup(CID,'mdsupdrs',anxious,row[6].value)
		sqlup(CID,'mdsupdrs',apathy,row[7].value)
		sqlup(CID,'mdsupdrs',dds,row[8].value)

		sqlup(CID,'mdsupdrs',sleep,row[10].value)
		sqlup(CID,'mdsupdrs',day_sleep,row[11].value)
		sqlup(CID,'mdsupdrs',pain,row[12].value)
		sqlup(CID,'mdsupdrs',urine,row[13].value)
		sqlup(CID,'mdsupdrs',constip,row[14].value)
		sqlup(CID,'mdsupdrs',light_head,row[15].value)
		sqlup(CID,'mdsupdrs',fatigue,row[16].value)

		sqlup(CID,'mdsupdrs',spch,row[17].value)
		sqlup(CID,'mdsupdrs',saliva,row[18].value)
		sqlup(CID,'mdsupdrs',swallow,row[19].value)
		sqlup(CID,'mdsupdrs',eating,row[20].value)
		sqlup(CID,'mdsupdrs',dress,row[21].value)
		sqlup(CID,'mdsupdrs',hygiene,row[22].value)
		sqlup(CID,'mdsupdrs',handwr,row[23].value)
		sqlup(CID,'mdsupdrs',hobby,row[24].value)
		sqlup(CID,'mdsupdrs',turn,row[25].value)
		sqlup(CID,'mdsupdrs',trem,row[26].value)
		sqlup(CID,'mdsupdrs',out_bed,row[27].value)
		sqlup(CID,'mdsupdrs',walk,row[28].value)
		sqlup(CID,'mdsupdrs',freeze,row[29].value)
		sqlup(CID,'mdsupdrs',taking_pdmeds,row[30].value)

		# State
		if row[31].value.lower() == "off":
			val = 'yes'
		elif row[31].value.lower() == 'on':
			val = 'no'
		elif not row[31].value:
			val = 'na'

		sqlup(CID,'mdsupdrs',offstate,val)

		sqlup(CID,'mdsupdrs',spch_motor,row[32].value)
		sqlup(CID,'mdsupdrs',face_exp,row[33].value)
		sqlup(CID,'mdsupdrs',rig_neck,row[34].value)
		sqlup(CID,'mdsupdrs',rig_rue,row[35].value)
		sqlup(CID,'mdsupdrs',rig_lue,row[36].value)
		sqlup(CID,'mdsupdrs',rig_rle,row[37].value)
		sqlup(CID,'mdsupdrs',rig_lle,row[38].value)

		sqlup(CID,'mdsupdrs',fngtp_r,row[39].value)
		sqlup(CID,'mdsupdrs',fngtp_l,row[40].value)
		sqlup(CID,'mdsupdrs',hndmv_r,row[41].value)
		sqlup(CID,'mdsupdrs',hndmv_l,row[42].value)
		sqlup(CID,'mdsupdrs',pro_r,row[43].value)
		sqlup(CID,'mdsupdrs',pro_l,row[44].value)
		sqlup(CID,'mdsupdrs',toe_r,row[45].value)
		sqlup(CID,'mdsupdrs',toe_l,row[46].value)
		sqlup(CID,'mdsupdrs',leg_r,row[47].value)
		sqlup(CID,'mdsupdrs',leg_l,row[48].value)
		sqlup(CID,'mdsupdrs',chair,row[49].value)
		sqlup(CID,'mdsupdrs',gait,row[50].value)
		sqlup(CID,'mdsupdrs',freez,row[51].value)
		sqlup(CID,'mdsupdrs',post_stab,row[52].value)
		sqlup(CID,'mdsupdrs',posture,row[53].value)
		sqlup(CID,'mdsupdrs',mvmnt,row[54].value)
		sqlup(CID,'mdsupdrs',postrem_r,row[55].value)
		sqlup(CID,'mdsupdrs',postrem_l,row[56].value)
		sqlup(CID,'mdsupdrs',kintrem_r,row[57].value)
		sqlup(CID,'mdsupdrs',kintrem_l,row[58].value)
		sqlup(CID,'mdsupdrs',trem_rue,row[59].value)
		sqlup(CID,'mdsupdrs',trem_lue,row[60].value)
		sqlup(CID,'mdsupdrs',trem_rle,row[61].value)
		sqlup(CID,'mdsupdrs',trem_lle,row[62].value)
		sqlup(CID,'mdsupdrs',trem_lip,row[63].value)
		sqlup(CID,'mdsupdrs',dys_imp,row[65].value)



		sqlup(CID,'mdsupdrs',total_one,row[82].value)
		sqlup(CID,'mdsupdrs',total_two,row[83].value)
		sqlup(CID,'mdsupdrs',total_three,row[84].value)
		sqlup(CID,'mdsupdrs',total_four,row[85].value)
